import json
from pathlib import Path

from openpyxl import load_workbook

wb_path = Path(__file__).resolve().parent.parent / "end-to-end (2).xlsx"
wb = load_workbook(wb_path, data_only=True)
ws = wb["Module1"]
headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(cell is not None and str(cell).strip() for cell in row):
        continue
    rows.append({headers[i]: row[i] for i in range(len(headers))})

print(json.dumps(rows, indent=2, ensure_ascii=False))
